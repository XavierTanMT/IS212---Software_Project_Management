from datetime import datetime, timezone, timedelta
from flask import request, jsonify, send_file
from firebase_admin import firestore
from google.cloud.firestore_v1.base_query import FieldFilter
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib.enums import TA_CENTER, TA_LEFT
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
import io
import csv

from . import reports_bp

def _viewer_id():
    """Get the current user ID from request headers"""
    return (request.headers.get("X-User-Id") or request.args.get("viewer_id") or "").strip()

def _is_admin_or_hr(db, user_id):
    """Check if user is admin or HR"""
    if not user_id:
        return False
    user_doc = db.collection("users").document(user_id).get()
    if not user_doc.exists:
        return False
    user_role = (user_doc.to_dict() or {}).get("role", "").lower()
    return user_role in ["admin", "hr"]

def parse_date(date_str):
    """Parse ISO date string to datetime - always returns timezone-aware datetime"""
    if not date_str:
        return None
    try:
        dt = datetime.fromisoformat(date_str.replace("Z", "+00:00"))
        # If datetime is naive, make it UTC aware
        if dt.tzinfo is None:
            dt = dt.replace(tzinfo=timezone.utc)
        return dt
    except Exception:
        # Fallback: try parsing with strptime for additional formats
        try:
            # Try parsing without microseconds
            dt = datetime.strptime(date_str[:19], '%Y-%m-%dT%H:%M:%S')
            return dt.replace(tzinfo=timezone.utc)
        except Exception:
            return None

def safe_get_user_info(user_data, field, default=""):
    """Safely extract user info from various data structures"""
    # If it's a list, take the first item (empty lists are falsy but check type first)
    if isinstance(user_data, list):
        if len(user_data) > 0:
            user_data = user_data[0]
        else:
            return default
    
    # Check if data is empty/falsy after list processing
    if not user_data:
        return default
    
    # If it's a dict, get the field
    if isinstance(user_data, dict):
        return user_data.get(field, default)
    
    return default

@reports_bp.route("/task-completion", methods=["GET"])
def task_completion_report():
    """
    Generate task completion report
    Query params:
    - format: pdf, csv, or xlsx (default: pdf)
    - user_id: filter by user (optional)
    - project_id: filter by project (optional)
    - start_date: filter by due date >= start_date (ISO format)
    - end_date: filter by due date <= end_date (ISO format)
    - report_type: user, project, or summary (default: summary)
    """
    db = firestore.client()
    viewer = _viewer_id()
    
    # Check if user is admin/HR
    if not _is_admin_or_hr(db, viewer):
        return jsonify({"error": "Unauthorized - Admin/HR only"}), 403
    
    # Get query parameters
    format_type = request.args.get("format", "pdf").lower()
    filter_user_id = request.args.get("user_id", "").strip()
    filter_project_id = request.args.get("project_id", "").strip()
    start_date_str = request.args.get("start_date", "").strip()
    end_date_str = request.args.get("end_date", "").strip()
    report_type = request.args.get("report_type", "summary").lower()
    
    # Parse dates
    start_date = parse_date(start_date_str)
    end_date = parse_date(end_date_str)
    
    # Build query
    query = db.collection("tasks")
    
    # Apply filters
    filters_applied = []
    if filter_user_id:
        query = query.where(filter=FieldFilter("assigned_to.user_id", "==", filter_user_id))
        filters_applied.append(f"User: {filter_user_id}")
    
    if filter_project_id:
        query = query.where(filter=FieldFilter("project_id", "==", filter_project_id))
        filters_applied.append(f"Project: {filter_project_id}")
    
    # Fetch tasks
    tasks = list(query.stream())
    
    # Filter by date range (post-query since Firestore has limited query capabilities)
    filtered_tasks = []
    for task_doc in tasks:
        task_data = task_doc.to_dict() or {}
        due_date_str = task_data.get("due_date")
        
        if due_date_str:
            due_date = parse_date(due_date_str)
            if due_date:
                # Check date range
                if start_date and due_date < start_date:
                    continue
                if end_date and due_date > end_date:
                    continue
        
        # Safely extract user info (handle dict, list, or None)
        assigned_to_data = task_data.get("assigned_to")
        created_by_data = task_data.get("created_by")
        
        filtered_tasks.append({
            "task_id": task_doc.id,
            "title": task_data.get("title", "N/A"),
            "status": task_data.get("status", "To Do"),
            "priority": task_data.get("priority", "Medium"),
            "due_date": due_date_str,
            "assigned_to": safe_get_user_info(assigned_to_data, "name", "Unassigned"),
            "assigned_to_id": safe_get_user_info(assigned_to_data, "user_id", ""),
            "project_id": task_data.get("project_id", ""),
            "created_by": safe_get_user_info(created_by_data, "name", "Unknown"),
            "created_at": task_data.get("created_at", ""),
        })
    
    if start_date:
        filters_applied.append(f"Start Date: {start_date_str}")
    if end_date:
        filters_applied.append(f"End Date: {end_date_str}")
    
    # Calculate statistics
    total_tasks = len(filtered_tasks)
    completed_tasks = sum(1 for t in filtered_tasks if t["status"] == "Completed")
    in_progress_tasks = sum(1 for t in filtered_tasks if t["status"] == "In Progress")
    todo_tasks = sum(1 for t in filtered_tasks if t["status"] == "To Do")
    blocked_tasks = sum(1 for t in filtered_tasks if t["status"] == "Blocked")
    
    completion_rate = (completed_tasks / total_tasks * 100) if total_tasks > 0 else 0
    
    stats = {
        "total_tasks": total_tasks,
        "completed": completed_tasks,
        "in_progress": in_progress_tasks,
        "todo": todo_tasks,
        "blocked": blocked_tasks,
        "completion_rate": round(completion_rate, 2)
    }
    
    # Generate report based on format
    if format_type == "pdf":
        return generate_pdf_report(filtered_tasks, stats, filters_applied, report_type)
    elif format_type == "csv":
        return generate_csv_report(filtered_tasks, stats, filters_applied)
    elif format_type == "xlsx":
        return generate_xlsx_report(filtered_tasks, stats, filters_applied, report_type)
    else:
        return jsonify({"error": "Invalid format. Use pdf, csv, or xlsx"}), 400


def generate_pdf_report(tasks, stats, filters, report_type):
    """Generate PDF report"""
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter, topMargin=0.75*inch, bottomMargin=0.75*inch)
    
    elements = []
    styles = getSampleStyleSheet()
    
    # Title
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=24,
        textColor=colors.HexColor('#667eea'),
        spaceAfter=30,
        alignment=TA_CENTER
    )
    elements.append(Paragraph("Task Completion Report", title_style))
    elements.append(Spacer(1, 0.2*inch))
    
    # Metadata
    meta_style = styles['Normal']
    elements.append(Paragraph(f"<b>Generated:</b> {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", meta_style))
    elements.append(Paragraph(f"<b>Report Type:</b> {report_type.capitalize()}", meta_style))
    
    if filters:
        elements.append(Paragraph(f"<b>Filters:</b> {', '.join(filters)}", meta_style))
    
    elements.append(Spacer(1, 0.3*inch))
    
    # Statistics Summary
    elements.append(Paragraph("<b>Summary Statistics</b>", styles['Heading2']))
    elements.append(Spacer(1, 0.1*inch))
    
    summary_data = [
        ['Metric', 'Value'],
        ['Total Tasks', str(stats['total_tasks'])],
        ['Completed', str(stats['completed'])],
        ['In Progress', str(stats['in_progress'])],
        ['To Do', str(stats['todo'])],
        ['Blocked', str(stats['blocked'])],
        ['Completion Rate', f"{stats['completion_rate']}%"],
    ]
    
    summary_table = Table(summary_data, colWidths=[3*inch, 2*inch])
    summary_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#667eea')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
    ]))
    elements.append(summary_table)
    elements.append(Spacer(1, 0.3*inch))
    
    # Tasks table
    elements.append(Paragraph("<b>Task Details</b>", styles['Heading2']))
    elements.append(Spacer(1, 0.1*inch))
    
    if tasks:
        # Table headers
        task_data = [['Title', 'Status', 'Priority', 'Assignee', 'Due Date']]
        
        # Add task rows
        for task in tasks[:50]:  # Limit to 50 tasks for PDF
            task_data.append([
                task['title'][:30] + '...' if len(task['title']) > 30 else task['title'],
                task['status'],
                task['priority'],
                task['assigned_to'][:20] + '...' if len(task['assigned_to']) > 20 else task['assigned_to'],
                task['due_date'][:10] if task['due_date'] else 'N/A'
            ])
        
        task_table = Table(task_data, colWidths=[2.5*inch, 1*inch, 0.8*inch, 1.5*inch, 1*inch])
        task_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#764ba2')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
            ('BACKGROUND', (0, 1), (-1, -1), colors.white),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
        ]))
        elements.append(task_table)
        
        if len(tasks) > 50:
            elements.append(Spacer(1, 0.2*inch))
            elements.append(Paragraph(f"<i>Note: Showing first 50 of {len(tasks)} tasks. Use CSV/XLSX for full data.</i>", styles['Normal']))
    else:
        elements.append(Paragraph("No tasks found matching the criteria.", styles['Normal']))
    
    # Build PDF
    doc.build(elements)
    buffer.seek(0)
    
    return send_file(
        buffer,
        mimetype='application/pdf',
        as_attachment=True,
        download_name=f'task_report_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf'
    )


def generate_csv_report(tasks, stats, filters):
    """Generate CSV report"""
    output = io.StringIO()
    writer = csv.writer(output)
    
    # Metadata
    writer.writerow(['Task Completion Report'])
    writer.writerow(['Generated:', datetime.now().strftime('%Y-%m-%d %H:%M:%S')])
    if filters:
        writer.writerow(['Filters:', ', '.join(filters)])
    writer.writerow([])
    
    # Summary statistics
    writer.writerow(['Summary Statistics'])
    writer.writerow(['Total Tasks', stats['total_tasks']])
    writer.writerow(['Completed', stats['completed']])
    writer.writerow(['In Progress', stats['in_progress']])
    writer.writerow(['To Do', stats['todo']])
    writer.writerow(['Blocked', stats['blocked']])
    writer.writerow(['Completion Rate', f"{stats['completion_rate']}%"])
    writer.writerow([])
    
    # Task details
    writer.writerow(['Task Details'])
    writer.writerow(['Task ID', 'Title', 'Status', 'Priority', 'Assignee', 'Project ID', 'Due Date', 'Created By', 'Created At'])
    
    for task in tasks:
        writer.writerow([
            task['task_id'],
            task['title'],
            task['status'],
            task['priority'],
            task['assigned_to'],
            task['project_id'],
            task['due_date'][:10] if task['due_date'] else 'N/A',
            task['created_by'],
            task['created_at'][:10] if task['created_at'] else 'N/A'
        ])
    
    output.seek(0)
    return send_file(
        io.BytesIO(output.getvalue().encode('utf-8')),
        mimetype='text/csv',
        as_attachment=True,
        download_name=f'task_report_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv'
    )


def generate_xlsx_report(tasks, stats, filters, report_type):
    """Generate Excel (XLSX) report"""
    workbook = openpyxl.Workbook()
    
    # Summary sheet
    summary_sheet = workbook.active
    summary_sheet.title = "Summary"
    
    # Header styling
    header_fill = PatternFill(start_color="667eea", end_color="667eea", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=14)
    
    # Title
    summary_sheet['A1'] = 'Task Completion Report'
    summary_sheet['A1'].font = Font(bold=True, size=18, color="667eea")
    summary_sheet.merge_cells('A1:B1')
    
    # Metadata
    summary_sheet['A3'] = 'Generated:'
    summary_sheet['B3'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    summary_sheet['A4'] = 'Report Type:'
    summary_sheet['B4'] = report_type.capitalize()
    
    if filters:
        summary_sheet['A5'] = 'Filters:'
        summary_sheet['B5'] = ', '.join(filters)
    
    # Statistics
    row = 7
    summary_sheet[f'A{row}'] = 'Summary Statistics'
    summary_sheet[f'A{row}'].font = Font(bold=True, size=12)
    
    row += 1
    summary_sheet[f'A{row}'] = 'Metric'
    summary_sheet[f'B{row}'] = 'Value'
    summary_sheet[f'A{row}'].fill = header_fill
    summary_sheet[f'B{row}'].fill = header_fill
    summary_sheet[f'A{row}'].font = header_font
    summary_sheet[f'B{row}'].font = header_font
    
    stats_data = [
        ('Total Tasks', stats['total_tasks']),
        ('Completed', stats['completed']),
        ('In Progress', stats['in_progress']),
        ('To Do', stats['todo']),
        ('Blocked', stats['blocked']),
        ('Completion Rate', f"{stats['completion_rate']}%"),
    ]
    
    for metric, value in stats_data:
        row += 1
        summary_sheet[f'A{row}'] = metric
        summary_sheet[f'B{row}'] = value
    
    # Column widths
    summary_sheet.column_dimensions['A'].width = 25
    summary_sheet.column_dimensions['B'].width = 20
    
    # Tasks sheet
    tasks_sheet = workbook.create_sheet("Tasks")
    
    # Headers
    headers = ['Task ID', 'Title', 'Status', 'Priority', 'Assignee', 'Project ID', 'Due Date', 'Created By', 'Created At']
    for col, header in enumerate(headers, start=1):
        cell = tasks_sheet.cell(row=1, column=col)
        cell.value = header
        cell.fill = PatternFill(start_color="764ba2", end_color="764ba2", fill_type="solid")
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal='center')
    
    # Task data
    for row_idx, task in enumerate(tasks, start=2):
        tasks_sheet.cell(row=row_idx, column=1).value = task['task_id']
        tasks_sheet.cell(row=row_idx, column=2).value = task['title']
        tasks_sheet.cell(row=row_idx, column=3).value = task['status']
        tasks_sheet.cell(row=row_idx, column=4).value = task['priority']
        tasks_sheet.cell(row=row_idx, column=5).value = task['assigned_to']
        tasks_sheet.cell(row=row_idx, column=6).value = task['project_id']
        tasks_sheet.cell(row=row_idx, column=7).value = task['due_date'][:10] if task['due_date'] else 'N/A'
        tasks_sheet.cell(row=row_idx, column=8).value = task['created_by']
        tasks_sheet.cell(row=row_idx, column=9).value = task['created_at'][:10] if task['created_at'] else 'N/A'
    
    # Auto-fit columns
    for col in range(1, 10):
        tasks_sheet.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 20
    
    # Save to buffer
    buffer = io.BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    
    return send_file(
        buffer,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'task_report_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    )


@reports_bp.route("/weekly-summary", methods=["GET"])
def weekly_summary_report():
    """
    Generate weekly team summary report
    Query params:
    - format: pdf, csv, or xlsx
    - week_start: ISO date string for week start (default: current week)
    """
    db = firestore.client()
    viewer = _viewer_id()
    
    if not _is_admin_or_hr(db, viewer):
        return jsonify({"error": "Unauthorized - Admin/HR only"}), 403
    
    # Calculate week range
    week_start_str = request.args.get("week_start", "").strip()
    if week_start_str:
        week_start = parse_date(week_start_str)
        if week_start is None:
            return jsonify({"error": "Invalid week_start date format"}), 400
    else:
        # Default to current week (Monday)
        today = datetime.now(timezone.utc)
        week_start = today - timedelta(days=today.weekday())
    
    week_end = week_start + timedelta(days=7)
    
    # This would query tasks created/updated in the week
    # For now, return a simplified version
    return jsonify({
        "message": "Weekly summary feature - coming soon",
        "week_start": week_start.isoformat(),
        "week_end": week_end.isoformat()
    })
